# coding=UTF-8
import openpyxl


def write_excel_xlsx(path, sheet_name, value):
    index = len(value)
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    for i in range(0, index):
        for j in range(0, len(value[i])):
            sheet.cell(row=i + 1, column=j + 1, value=str(value[i][j]))
    workbook.save(path)
    print("xlsx格式表格写入数据成功！")


def read_excel_xlsx(path, sheet_name):
    workbook = openpyxl.load_workbook(path)
    # sheet = wb.get_sheet_by_name(sheet_name)这种方式已经弃用，不建议使用
    sheet = workbook[sheet_name]
    for row in sheet.rows:
        for cell in row:
            print(cell.value, "\t", end="")
        print()


book_name_xlsx = 'xlsx格式测试工作簿.xlsx'

sheet_name_xlsx = 'xlsx格式测试表'

value3 = [["姓名", "性别", "年龄", "城市", "职业"],
          ["111", "女", "66", "石家庄", "运维工程师"],
          ["222", "男", "55", "南京", "饭店老板"],
          ["333", "女", "27", "苏州", "保安"], ]

write_excel_xlsx(book_name_xlsx, sheet_name_xlsx, value3)
read_excel_xlsx(book_name_xlsx, sheet_name_xlsx)
list3=[]
list1=["姓名", "性别", "年龄", "城市", "职业"]
list2=["111", "女", "66", "石家庄", "运维工程师"]
list3.append(list2)
list3.append(list1)
print(list3)

list4=[]
name="sss"
age="xxxx"
list4.append(name)
list4.append(age)
print(list4)
