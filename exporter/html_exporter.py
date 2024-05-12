import datetime
import json
import shutil
import threading
import time
import openpyxl
import xmltodict

from entity.contact import Contact
from exporter.avatar_exporter import AvatarExporter
from exporter.emoji_exporter import EmojiExporter
from exporter.image_exporter import ImageExporter
from exporter.video_exporter import VideoExporter
from log import LOG
from entity.moment_msg import MomentMsg
from pathlib import Path

book_name_xlsx = '数据.xlsx'

sheet_name_xlsx = '数据'



def write_excel_xlsx(path, sheet_name, value):
    index = len(value)
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    for i in range(0, index):
        for j in range(0, len(value[i])):
            sheet.cell(row=i + 1, column=j + 1, value=str(value[i][j]))
    workbook.save(path)
    print("xlsx格式表格写入数据成功！")


def read_excel_xlsx(path, sheet_name):
    workbook = openpyxl.load_workbook(path)
    # sheet = wb.get_sheet_by_name(sheet_name)这种方式已经弃用，不建议使用
    sheet = workbook[sheet_name]
    for row in sheet.rows:
        for cell in row:
            print(cell.value, "\t", end="")
        print()

def get_img_div_css(size: int) -> str:
    if size == 1:
        return 'width:10rem; overflow:hidden'
    else:
        return 'width:19rem; overflow:hidden'


def get_img_css(size: int) -> str:
    """object-fit: cover; 预览图居中裁剪
       cursor:pointer; 手形鼠标
    """
    img_style = "object-fit:cover;cursor:pointer;"
    if size == 1:
        return f'width:10rem;height:10rem;{img_style}'
    elif size == 2:
        return f'width:8rem;height:8rem;float:left;margin-bottom:0.2rem;margin-right:0.2rem;{img_style}'
    elif size == 4:
        return f'width:8rem;height:8rem;float:left;margin-bottom:0.2rem;margin-right:0.2rem;{img_style}'
    else:
        return f'width:5rem;height:5rem;float:left;margin-bottom:0.2rem;margin-right:0.2rem;{img_style}'


class HtmlExporter(threading.Thread):

    def __init__(self, gui: 'Gui', dir_name: str, contacts_map: dict[str, Contact], begin_date: datetime.date,
                 end_date: datetime.date, download_pic: int, convert_video: int):
        self.dir_name = dir_name
        if Path(f"output/{self.dir_name}").exists():
            shutil.rmtree(f"output/{self.dir_name}")
        shutil.copytree("resource/template/", f"output/{self.dir_name}")

        self.gui = gui
        self.avatar_exporter = AvatarExporter(dir_name)
        self.image_exporter = ImageExporter(dir_name)
        self.video_exporter = VideoExporter(dir_name)
        self.html_head = None
        self.html_end = None
        self.file = None
        self.contacts_map = contacts_map
        self.begin_date = begin_date
        self.end_date = end_date
        self.download_pic = download_pic
        self.convert_video = convert_video
        self.stop_flag = False
        super().__init__()

    def run(self) -> None:

        with open(f"resource/template.html", encoding='utf-8') as f:
            content = f.read()
            self.html_head, self.html_end = content.split('/*内容分割线*/')
        self.file = open(f"output/{self.dir_name}/index.html", 'w', encoding='utf-8')

        if self.gui.account_info and self.gui.account_info.get('wxid'):
            self.avatar_exporter.get_avatar_path(self.gui.account_info.get('wxid'))
            self.html_head = self.html_head.replace("{my_wxid}", f"{self.gui.account_info.get('wxid')}")
            from app.DataBase import micro_msg_db
            my_info = micro_msg_db.get_contact_by_username(self.gui.account_info.get('wxid'))
            self.html_head = self.html_head.replace("{my_name}", f"{my_info[4]}")

        from app.DataBase import sns_db
        cover_url = sns_db.get_cover_url()
        if cover_url:
            cover_path = self.image_exporter.save_image(cover_url, 'image')
            self.html_head = self.html_head.replace("{cover_path}", cover_path)

        self.file.write(self.html_head)
        # 加一天
        end_date = self.end_date + datetime.timedelta(days=1)
        begin_time = time.mktime(datetime.datetime(self.begin_date.year, self.begin_date.month, self.begin_date.day).timetuple())
        end_time = time.mktime(datetime.datetime(end_date.year, end_date.month, end_date.day).timetuple())

        self.gui.video_decrypter.decrypt_videos(self, self.begin_date, end_date, self.dir_name, self.convert_video)


        message_datas = sns_db.get_messages_in_time(begin_time, end_time)

        # 写入excel的
        listhead = ['微信ID','备注或用户名','信息','图文链接或视频链接','链接说明']

        resultlist = []
        resultlist.append(listhead)

        for index, message_data in enumerate(message_datas):
            if not self.stop_flag:
                if message_data[0] in self.contacts_map:
                    self.export_msg(message_data[1], self.contacts_map, self.download_pic,resultlist)
                    # 更新进度条 前30%视频处理  后70%其他处理
                    progress = round(index / len(message_datas) * 70)
                    self.gui.update_export_progressbar(30 + progress)

        # 写入表格
        write_excel_xlsx(f"output/{self.dir_name}/" + book_name_xlsx, sheet_name_xlsx, resultlist)

        self.gui.update_export_progressbar(100)
        self.finish_file()
        self.gui.export_succeed()

    def stop(self) -> None:
        self.stop_flag = True

    def export_msg(self, message: str, contacts_map: dict[str, Contact], download_pic: int,resultlist) -> None:

        LOG.info(message)
        # force_list: 强制要求转media为list
        msg_dict = xmltodict.parse(message, force_list={'media'})
        msg_json = json.dumps(msg_dict)
        msg = MomentMsg.from_json(msg_json)

        # 微信ID
        username = msg.timelineObject.username
        # 头像路径
        avatar_path = self.avatar_exporter.get_avatar_path(username)

        contact = contacts_map.get(username)
        #print(contact)
        # 备注， 或用户名
        remark = contact.remark if contact.remark else contact.nickName
        infolist = []
        infolist.append(username)
        #infolist.append(contact.exTraBuf.encode('utf-8'))
        infolist.append(remark)

        # 朋友圈图片
        images = self.image_exporter.get_images(msg, download_pic)

        # 朋友圈视频
        videos = self.video_exporter.get_videos(msg)

        # 样式   3:链接样式
        content_style = msg.timelineObject.ContentObject.contentStyle

        html = ' <div class="row item">\n'
        html += '    <div class="col-xs-2">\n'
        html += '         <div class="logo01_box">\n'
        html += f'             <img src="{avatar_path}" />\n'
        html += '         </div>\n'
        html += '    </div>\n'
        html += '    <div class="col-xs-10 xs8">\n'
        html += '          <div class="towp">\n'
        html += f'              <p class="p1">{remark}</p>\n'
        if msg.timelineObject.contentDesc:
            content_desc = msg.timelineObject.contentDesc.replace("\n", "<br>")
            content_desc = EmojiExporter.replace_emoji(content_desc)
            html += f'              <p class="p2">{content_desc}</p>\n'
            # 输出到表格
            infolist.append(content_desc)
        else:
            infolist.append("信息为空")
        html += '          </div>\n'

        # 超链接
        if content_style == 3:
            html += f'  <a href="{msg.timelineObject.ContentObject.contentUrl}" target="_blank">\n'
            html += '      <div class ="out_link" >\n'
            if images:
                thumb_path, image_path = images[0]
                html += f'        <img src = "{thumb_path}"/>\n'
            html += f'       <div class ="text" >{msg.timelineObject.ContentObject.title}</div>\n'
            html += '      </div >\n'
            html += '   </a>\n'

            #输出到表格
            infolist.append(msg.timelineObject.ContentObject.contentUrl)
            infolist.append(msg.timelineObject.ContentObject.title)
        # 视频号
        elif msg.timelineObject.ContentObject.finderFeed:
            html += f'     <div style="width:10rem; overflow:hidden">\n'
            # 视频号图片
            thumb_path = self.image_exporter.get_finder_images(msg)
            html += f'       <img src="{thumb_path}" onclick="openWarningOverlay(event)" style="width:10rem;height:10rem;object-fit:cover;cursor:pointer;"/>\n'
            html += '      </div>\n'

            # 视频号说明
            html += '          <div class="texts_box">\n'
            nickname = msg.timelineObject.ContentObject.finderFeed.nickname
            desc = msg.timelineObject.ContentObject.finderFeed.desc
            html += f'            <p class="location">视频号 · {nickname} · {desc}</p>\n'
            html += '         </div>\n'

            # 输出到表格
            infolist.append("thumb_path")
            infolist.append("视频号" + nickname  + desc)
        # 普通朋友圈
        else:
            html += f'     <div style="{get_img_div_css(len(images))}">\n'
            for thumb_path, image_path in images:
                html += f'     <img src="{thumb_path}" full_img="{image_path}" onclick="openFullSize(event)" style="{get_img_css(len(images))}"/>\n'
            html += '      </div>\n'

            html += '      <div>\n'
            for video_path in videos:
                html += f'     <video controls height="500">\n'
                html += f'        <source  src="{video_path}" type="video/mp4">\n'
                html += f'     <video>\n'
            html += '      </div>\n'
            infolist.append(" ")
            infolist.append(" ")


        html += '          <div class="texts_box">\n'
        if msg.timelineObject.location and msg.timelineObject.location.poiName:
            html += f'        <p class="location">{msg.timelineObject.location.poiName}</p>\n'
        html += f'            <p class="time">{msg.timelineObject.create_time}</p>\n'
        html += '         </div>\n'
        html += '    </div>\n'
        html += '</div>\n'

        resultlist.append(infolist)
        self.file.write(html)

    def finish_file(self):
        self.file.write(self.html_end)
        self.file.close()
